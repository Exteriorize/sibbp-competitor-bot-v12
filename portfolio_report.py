from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from competitor_service import flatten_snapshot_items, get_manual_review_priority_rows, get_portfolio_priority_rows, summarize_all_competitors
from lifecycle_store import get_archive_items, get_recent_changes


RUB_FORMAT = '#,##0 "₽"'
RUB_M2_FORMAT = '#,##0.00 "₽/м²"'
AREA_FORMAT = '#,##0.0 "м²"'


def _apply_header_style(ws):
    fill = PatternFill(fill_type="solid", fgColor="D9EAD3")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = fill


def _autowidth(ws):
    for column_cells in ws.columns:
        max_len = 0
        col_idx = column_cells[0].column
        col_letter = get_column_letter(col_idx)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_len:
                max_len = len(value)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 14), 44)


def _style_sheet(ws):
    _apply_header_style(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    _autowidth(ws)


def _fill_number_formats(ws, columns: Dict[str, str]):
    header_map = {str(ws.cell(row=1, column=col).value): col for col in range(1, ws.max_column + 1)}
    for header, fmt in columns.items():
        col = header_map.get(header)
        if not col:
            continue
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=col).number_format = fmt


def create_portfolio_report(snapshots: List[Dict], output_path: Optional[str] = None) -> str:
    if output_path is None:
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
        output_path = str(Path("reports") / f"portfolio_report_{timestamp}.xlsx")

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)

    stats = summarize_all_competitors(snapshots)
    items = flatten_snapshot_items(snapshots)
    changes = get_recent_changes(days=14)
    archive_items = get_archive_items()
    priority_rows = get_portfolio_priority_rows(snapshots)
    review_rows = get_manual_review_priority_rows()

    own_stats = stats.get("own_company", {})
    competitor_stats = stats.get("competitors", {})

    summary_rows = [
        {"Показатель": "Объектов в базе", "Значение": stats.get("competitors_total", 0)},
        {"Показатель": "Моя компания — помещений", "Значение": own_stats.get("count", 0)},
        {"Показатель": "Моя компания — площадь", "Значение": own_stats.get("total_area", 0)},
        {"Показатель": "Моя компания — средневзвешенная цена", "Значение": own_stats.get("avg_price", 0)},
        {"Показатель": "Моя компания — суммарная стоимость", "Значение": own_stats.get("total_price", 0)},
        {"Показатель": "Конкуренты — помещений", "Значение": competitor_stats.get("count", 0)},
        {"Показатель": "Конкуренты — площадь", "Значение": competitor_stats.get("total_area", 0)},
        {"Показатель": "Конкуренты — средневзвешенная цена", "Значение": competitor_stats.get("avg_price", 0)},
        {"Показатель": "Конкуренты — суммарная стоимость", "Значение": competitor_stats.get("total_price", 0)},
        {"Показатель": "Всего найдено помещений", "Значение": stats.get("count", 0)},
        {"Показатель": "Всего суммарная площадь", "Значение": stats.get("total_area", 0)},
        {"Показатель": "Всего средневзвешенная цена", "Значение": stats.get("avg_price", 0)},
        {"Показатель": "Всего суммарная стоимость", "Значение": stats.get("total_price", 0)},
        {"Показатель": "Неподтвержденных объектов", "Значение": stats.get("unconfirmed_count", 0)},
        {"Показатель": "Выбывших объектов в архиве", "Значение": stats.get("removed_count", 0)},
        {"Показатель": "Ручных помещений к проверке", "Значение": stats.get("review_items_count", 0)},
    ]
    df_summary = pd.DataFrame(summary_rows)

    competitor_rows = []
    for snapshot in snapshots:
        competitor = snapshot.get("competitor", {})
        stats_row = snapshot.get("stats", {})
        latest_record = snapshot.get("latest_record") or {}
        freshness = snapshot.get("freshness", {})
        lifecycle = snapshot.get("lifecycle", {})
        competitor_rows.append(
            {
                "Группа": "Моя компания" if competitor.get("entity_role") == "own_company" else "Конкурент",
                "Название": competitor.get("name", ""),
                "Режим": "Парсинг сайта" if competitor.get("mode") == "parsed" else "Ручной учет",
                "Свободных помещений": stats_row.get("count", 0),
                "Свободная площадь, м²": stats_row.get("total_area", 0),
                "Средневзвешенная цена, ₽/м²": stats_row.get("avg_price", 0),
                "Суммарная стоимость, ₽": stats_row.get("total_price", 0),
                "Неподтверждено": lifecycle.get("unconfirmed_count", 0),
                "В архиве": lifecycle.get("removed_count", 0),
                "К проверке вручную": snapshot.get("review_items_count", 0),
                "Свежесть данных": freshness.get("freshness_label", ""),
                "Последняя проверка": freshness.get("last_checked_at", ""),
                "Приоритет прозвона": snapshot.get("priority_label", "Низкий"),
                "Причины приоритета": "; ".join(snapshot.get("priority_reasons") or []) or "—",
                "Источник": latest_record.get("source_label", "") or ("Сайт" if competitor.get("mode") == "parsed" else ""),
                "Достоверность": latest_record.get("reliability_label", "") or ("Высокая" if competitor.get("mode") == "parsed" else ""),
                "Комментарий": latest_record.get("comment", ""),
                "Ошибка": snapshot.get("error", ""),
            }
        )
    df_competitors = pd.DataFrame(competitor_rows)

    item_rows = []
    for item in items:
        item_rows.append(
            {
                "Группа": "Моя компания" if item.get("entity_role") == "own_company" else "Конкурент",
                "Компания": item.get("company", ""),
                "Тип": item.get("type", ""),
                "Название": item.get("title", ""),
                "Площадь, м²": item.get("area"),
                "Ставка за м², ₽": item.get("price_value") or item.get("price_per_sqm"),
                "Общая стоимость, ₽": item.get("total_price_value") or item.get("total_price"),
                "Источник": item.get("source_label", "") or item.get("source_kind", ""),
                "Достоверность": item.get("reliability_label", ""),
                "Ссылка": item.get("source_url") or item.get("url", ""),
            }
        )
    df_items = pd.DataFrame(item_rows)

    archive_rows = []
    for row in archive_items:
        archive_rows.append(
            {
                "Компания": row.get("competitor_name", ""),
                "Статус": row.get("status", ""),
                "Тип": row.get("type", ""),
                "Название": row.get("title", ""),
                "Площадь, м²": row.get("area", 0),
                "Ставка за м², ₽": row.get("price_per_sqm", 0),
                "Общая стоимость, ₽": row.get("total_price", 0),
                "Впервые найдено": row.get("first_seen", ""),
                "Последний раз найдено": row.get("last_seen", ""),
                "Ссылка": row.get("source_url", ""),
            }
        )
    df_archive = pd.DataFrame(archive_rows)

    change_rows = []
    for row in changes:
        change_rows.append(
            {
                "Дата": row.get("event_at", ""),
                "Компания": row.get("competitor_name", ""),
                "Тип": row.get("type", ""),
                "Название": row.get("title", ""),
                "Событие": row.get("event_type", ""),
                "Старое значение": row.get("old_value", ""),
                "Новое значение": row.get("new_value", ""),
                "Комментарий": row.get("note", ""),
            }
        )
    df_changes = pd.DataFrame(change_rows)
    df_priority = pd.DataFrame(priority_rows)
    df_review = pd.DataFrame(review_rows)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df_summary.to_excel(writer, index=False, sheet_name="Сводка")
        df_competitors.to_excel(writer, index=False, sheet_name="База")
        df_items.to_excel(writer, index=False, sheet_name="Помещения")
        df_changes.to_excel(writer, index=False, sheet_name="Изменения")
        df_archive.to_excel(writer, index=False, sheet_name="Архив")
        df_priority.to_excel(writer, index=False, sheet_name="Прозвон")
        df_review.to_excel(writer, index=False, sheet_name="Ручная ревизия")

        for name in writer.book.sheetnames:
            _style_sheet(writer.book[name])

        ws_summary = writer.book["Сводка"]
        for row in range(2, ws_summary.max_row + 1):
            label = ws_summary[f"A{row}"].value or ""
            cell = ws_summary[f"B{row}"]
            if "площадь" in str(label).lower():
                cell.number_format = AREA_FORMAT
            elif "цена" in str(label).lower():
                cell.number_format = RUB_M2_FORMAT
            elif "стоимость" in str(label).lower():
                cell.number_format = RUB_FORMAT

        _fill_number_formats(writer.book["База"], {
            "Свободная площадь, м²": AREA_FORMAT,
            "Средневзвешенная цена, ₽/м²": RUB_M2_FORMAT,
            "Суммарная стоимость, ₽": RUB_FORMAT,
        })
        _fill_number_formats(writer.book["Помещения"], {
            "Площадь, м²": AREA_FORMAT,
            "Ставка за м², ₽": RUB_M2_FORMAT,
            "Общая стоимость, ₽": RUB_FORMAT,
        })
        _fill_number_formats(writer.book["Архив"], {
            "Площадь, м²": AREA_FORMAT,
            "Ставка за м², ₽": RUB_M2_FORMAT,
            "Общая стоимость, ₽": RUB_FORMAT,
        })
        _fill_number_formats(writer.book["Прозвон"], {"Свободная площадь, м²": AREA_FORMAT})
        _fill_number_formats(writer.book["Ручная ревизия"], {
            "Площадь, м²": AREA_FORMAT,
            "Ставка за м², ₽": RUB_M2_FORMAT,
            "Общая стоимость, ₽": RUB_FORMAT,
        })

        for sheet_name, link_header in (("Помещения", "Ссылка"), ("Архив", "Ссылка"), ("Ручная ревизия", "Ссылка")):
            ws = writer.book[sheet_name]
            header_map = {str(ws.cell(row=1, column=col).value): col for col in range(1, ws.max_column + 1)}
            col = header_map.get(link_header)
            if col:
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=col)
                    if cell.value:
                        cell.hyperlink = str(cell.value)
                        cell.style = "Hyperlink"

    return output_path
