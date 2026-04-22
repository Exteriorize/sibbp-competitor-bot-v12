from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from analytics import summarize
from lifecycle_store import get_archive_items, get_recent_changes


RUB_FORMAT = '#,##0 "₽"'
RUB_M2_FORMAT = '#,##0.00 "₽/м²"'
AREA_FORMAT = '#,##0.0 "м²"'


def _apply_header_style(ws):
    fill = PatternFill(fill_type="solid", fgColor="D9EAD3")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = fill


def _autowidth(ws):
    for column_cells in ws.columns:
        max_len = 0
        col_idx = column_cells[0].column
        col_letter = get_column_letter(col_idx)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_len:
                max_len = len(value)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 14), 42)


def _style_sheet(ws):
    _apply_header_style(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    _autowidth(ws)


def create_report(items: List[Dict], competitor: Optional[Dict] = None, lifecycle: Optional[Dict] = None, output_path: Optional[str] = None) -> str:
    competitor = competitor or {}
    competitor_name = str(competitor.get("name") or competitor.get("short_name") or "competitor")
    competitor_code = str(competitor.get("code") or "")
    lifecycle = lifecycle or {}

    if output_path is None:
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        safe_name = "".join(ch for ch in competitor_name if ch.isalnum() or ch in ("_", "-", " ")).strip() or "competitor"
        output_path = str(Path("reports") / f"{safe_name}_report_{timestamp}.xlsx")

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)

    summary = summarize(items)
    archive_items = lifecycle.get("archive_items") or (get_archive_items(competitor_code) if competitor_code else [])
    changes = get_recent_changes(days=14, competitor_code=competitor_code) if competitor_code else []

    summary_rows = [
        {"Показатель": "Найдено помещений", "Значение": summary.get("count", 0)},
        {"Показатель": "Суммарная площадь", "Значение": summary.get("total_area", 0)},
        {"Показатель": "Средняя цена", "Значение": summary.get("avg_price", 0)},
        {"Показатель": "Суммарная стоимость", "Значение": summary.get("total_price", 0)},
        {"Показатель": "Неподтвержденных объектов", "Значение": lifecycle.get("unconfirmed_count", 0)},
        {"Показатель": "Выбывших объектов", "Значение": lifecycle.get("removed_count", 0)},
    ]
    if competitor_name:
        summary_rows.insert(0, {"Показатель": "Конкурент", "Значение": competitor_name})

    rows = []
    for item in items:
        total_price = item.get("total_price_value") or item.get("total_price") or 0
        area = item.get("area") or 0
        rate = item.get("price_value") or item.get("price_per_sqm") or 0
        if not total_price and area and rate:
            total_price = round(float(area) * float(rate), 2)
        rows.append(
            {
                "Тип": item.get("type", ""),
                "Название": item.get("title", ""),
                "Площадь, м²": area,
                "Ставка за м², ₽": rate,
                "Общая стоимость, ₽": total_price,
                "Источник": item.get("source_label", "") or item.get("source_kind", ""),
                "Достоверность": item.get("reliability_label", ""),
                "Ссылка": item.get("source_url") or item.get("url", ""),
            }
        )

    archive_rows = []
    for row in archive_items:
        archive_rows.append(
            {
                "Статус": row.get("status", ""),
                "Тип": row.get("type", ""),
                "Название": row.get("title", ""),
                "Площадь, м²": row.get("area", 0),
                "Ставка за м², ₽": row.get("price_per_sqm", 0),
                "Общая стоимость, ₽": row.get("total_price", 0),
                "Впервые найдено": row.get("first_seen", ""),
                "Последний раз найдено": row.get("last_seen", ""),
                "Ссылка": row.get("source_url", ""),
            }
        )

    change_rows = []
    for row in changes:
        change_rows.append(
            {
                "Дата": row.get("event_at", ""),
                "Тип": row.get("type", ""),
                "Название": row.get("title", ""),
                "Событие": row.get("event_type", ""),
                "Старое значение": row.get("old_value", ""),
                "Новое значение": row.get("new_value", ""),
                "Комментарий": row.get("note", ""),
            }
        )

    df_summary = pd.DataFrame(summary_rows, columns=["Показатель", "Значение"])
    df_items = pd.DataFrame(rows, columns=["Тип", "Название", "Площадь, м²", "Ставка за м², ₽", "Общая стоимость, ₽", "Источник", "Достоверность", "Ссылка"])
    df_archive = pd.DataFrame(archive_rows, columns=["Статус", "Тип", "Название", "Площадь, м²", "Ставка за м², ₽", "Общая стоимость, ₽", "Впервые найдено", "Последний раз найдено", "Ссылка"])
    df_changes = pd.DataFrame(change_rows, columns=["Дата", "Тип", "Название", "Событие", "Старое значение", "Новое значение", "Комментарий"])

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df_summary.to_excel(writer, index=False, sheet_name="Сводка")
        df_items.to_excel(writer, index=False, sheet_name="Помещения")
        df_archive.to_excel(writer, index=False, sheet_name="Архив")
        df_changes.to_excel(writer, index=False, sheet_name="Изменения")

        for name in ("Сводка", "Помещения", "Архив", "Изменения"):
            ws = writer.book[name]
            _style_sheet(ws)

        ws_summary = writer.book["Сводка"]
        for row in range(2, ws_summary.max_row + 1):
            label = ws_summary[f"A{row}"].value
            cell = ws_summary[f"B{row}"]
            if label == "Суммарная площадь":
                cell.number_format = AREA_FORMAT
            elif label == "Средняя цена":
                cell.number_format = RUB_M2_FORMAT
            elif label == "Суммарная стоимость":
                cell.number_format = RUB_FORMAT

        ws_items = writer.book["Помещения"]
        for row in range(2, ws_items.max_row + 1):
            ws_items[f"C{row}"].number_format = AREA_FORMAT
            ws_items[f"D{row}"].number_format = RUB_M2_FORMAT
            ws_items[f"E{row}"].number_format = RUB_FORMAT
            link_cell = ws_items[f"H{row}"]
            if link_cell.value:
                link_cell.hyperlink = str(link_cell.value)
                link_cell.style = "Hyperlink"

        ws_archive = writer.book["Архив"]
        if ws_archive.max_column >= 5:
            for row in range(2, ws_archive.max_row + 1):
                ws_archive[f"D{row}"].number_format = AREA_FORMAT
                ws_archive[f"E{row}"].number_format = RUB_M2_FORMAT
                ws_archive[f"F{row}"].number_format = RUB_FORMAT
                link_cell = ws_archive[f"I{row}"]
                if link_cell.value:
                    link_cell.hyperlink = str(link_cell.value)
                    link_cell.style = "Hyperlink"

    return output_path
