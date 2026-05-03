from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Iterable, Optional

import pandas as pd
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

AREA_FORMAT = '#,##0.0 "м²"'
RUB_FORMAT = '#,##0 "₽"'
RUB_M2_FORMAT = '#,##0.00 "₽/м²"'
INT_FORMAT = '#,##0'

CATEGORY_NAMES = {
    'commercial': 'Офисы / свободного назначения / торговые',
    'industrial': 'Производства / склады',
}


def _style_table(ws):
    fill = PatternFill(fill_type='solid', fgColor='D9EAD3')
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.fill = fill
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = max(len(str(c.value or '')) for c in col) + 2
        ws.column_dimensions[letter].width = min(max(width, 14), 44)
    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            h = str(ws.cell(1, col).value or '').lower()
            cell = ws.cell(row, col)
            if 'площад' in h or 'м²' in h:
                cell.number_format = AREA_FORMAT
            elif 'цена' in h or 'ставка' in h or '₽/м²' in h:
                cell.number_format = RUB_M2_FORMAT
            elif 'стоимость' in h or '₽' in h:
                cell.number_format = RUB_FORMAT
            elif 'помещ' in h or 'количество' in h:
                cell.number_format = INT_FORMAT


def _write_sheet(writer, df: pd.DataFrame, name: str):
    df.to_excel(writer, sheet_name=name[:31], index=False)
    _style_table(writer.book[name[:31]])


def _add_chart(writer, sheet: str, title: str, cols: list[str], y_title: str, pos: str, chart_sheet: str = 'Графики'):
    wb = writer.book
    ws = wb[sheet[:31]]
    cws = wb[chart_sheet[:31]] if chart_sheet[:31] in wb.sheetnames else wb.create_sheet(chart_sheet[:31])
    if cws['A1'].value is None:
        cws['A1'] = chart_sheet
        cws['A1'].font = Font(bold=True, size=16, color='FFFFFF')
        cws['A1'].fill = PatternFill(fill_type='solid', fgColor='1F4E78')
    headers = {str(ws.cell(1, c).value): c for c in range(1, ws.max_column + 1)}
    chart_cols = [headers[x] for x in cols if x in headers]
    if ws.max_row < 2 or not chart_cols:
        cws[pos] = 'Недостаточно данных для графика'
        return
    chart = LineChart()
    chart.title = title
    chart.y_axis.title = y_title
    chart.x_axis.title = 'Дата'
    chart.height = 11
    chart.width = 29
    chart.style = 13
    chart.legend.position = 'r'
    chart.dLbls = DataLabelList()
    chart.dLbls.showVal = True
    chart.dLbls.showCatName = False
    chart.dLbls.showSerName = False
    for c in chart_cols:
        chart.add_data(Reference(ws, min_col=c, min_row=1, max_row=ws.max_row), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=ws.max_row))
    cws.add_chart(chart, pos)


def _pivot(df: pd.DataFrame, value_col: str, suffix: str) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=['Дата'])
    out = df.pivot_table(index='snapshot_date', columns='competitor_name', values=value_col, aggfunc='last', fill_value=0).reset_index()
    out = out.rename(columns={'snapshot_date': 'Дата'})
    return out.rename(columns={c: f'{c} — {suffix}' for c in out.columns if c != 'Дата'})


def _base_output(prefix: str, output_path: Optional[str]) -> str:
    if output_path:
        return output_path
    Path('reports').mkdir(parents=True, exist_ok=True)
    return str(Path('reports') / f'{prefix}_{datetime.now().strftime("%Y-%m-%d_%H-%M")}.xlsx')


def _prepare_common(df: pd.DataFrame, cols: list[str]) -> pd.DataFrame:
    for col in cols:
        if col not in df.columns:
            df[col] = ''
    for col in ['count', 'total_area', 'avg_price', 'total_price']:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
    if 'competitor_name' in df.columns:
        df['competitor_name'] = df['competitor_name'].fillna('').replace('', 'Без названия')
    return df


def create_dynamics_report(records: Iterable[dict], competitor_name: str, output_path: Optional[str] = None) -> str:
    output_path = _base_output('dynamic_' + ''.join(ch for ch in competitor_name if ch.isalnum() or ch in ('_', '-', ' ')).strip(), output_path)
    df = _prepare_common(pd.DataFrame(list(records)), ['snapshot_date', 'count', 'total_area', 'avg_price', 'total_price', 'unconfirmed_count', 'removed_count', 'data_freshness', 'competitor_name'])
    df = df[['snapshot_date', 'count', 'total_area', 'avg_price', 'total_price', 'unconfirmed_count', 'removed_count', 'data_freshness', 'competitor_name']].rename(columns={
        'snapshot_date': 'Дата', 'count': 'Найдено помещений', 'total_area': 'Свободная площадь, м²', 'avg_price': 'Средняя цена, ₽/м²', 'total_price': 'Суммарная стоимость, ₽', 'unconfirmed_count': 'Неподтвержденных объектов', 'removed_count': 'Объектов в архиве', 'data_freshness': 'Свежесть данных', 'competitor_name': 'Конкурент'
    })
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        _write_sheet(writer, df, 'История')
        _add_chart(writer, 'История', f'Площадь — {competitor_name}', ['Свободная площадь, м²'], 'Площадь, м²', 'A3')
        _add_chart(writer, 'История', f'Помещения — {competitor_name}', ['Найдено помещений'], 'Количество', 'A27')
        _add_chart(writer, 'История', f'Ставка — {competitor_name}', ['Средняя цена, ₽/м²'], '₽/м²', 'A51')
    return output_path


def create_portfolio_dynamics_report(records: Iterable[dict], output_path: Optional[str] = None) -> str:
    output_path = _base_output('portfolio_dynamics', output_path)
    df = _prepare_common(pd.DataFrame(list(records)), ['snapshot_date', 'competitors_included', 'count', 'total_area', 'avg_price', 'total_price', 'unconfirmed_count', 'removed_count', 'data_freshness'])
    df = df.rename(columns={'snapshot_date': 'Дата', 'competitors_included': 'Объектов в расчете', 'count': 'Найдено помещений', 'total_area': 'Свободная площадь, м²', 'avg_price': 'Средняя цена, ₽/м²', 'total_price': 'Суммарная стоимость, ₽', 'unconfirmed_count': 'Неподтвержденных объектов', 'removed_count': 'Объектов в архиве', 'data_freshness': 'Свежесть данных'})
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        _write_sheet(writer, df, 'История')
        _add_chart(writer, 'История', 'Площадь — вся база', ['Свободная площадь, м²'], 'Площадь, м²', 'A3')
        _add_chart(writer, 'История', 'Помещения — вся база', ['Найдено помещений'], 'Количество', 'A27')
        _add_chart(writer, 'История', 'Ставка — вся база', ['Средняя цена, ₽/м²'], '₽/м²', 'A51')
    return output_path


def _category_df(category_records: list[dict], code: str) -> pd.DataFrame:
    df = pd.DataFrame(category_records)
    if df.empty:
        return pd.DataFrame(columns=['snapshot_date', 'competitor_name', 'count', 'total_area', 'avg_price', 'total_price'])
    df = _prepare_common(df, ['snapshot_date', 'competitor_name', 'category_code', 'count', 'total_area', 'avg_price', 'total_price'])
    return df.loc[df['category_code'] == code].copy()


def create_role_comparison_report(records: Iterable[dict], output_path: Optional[str] = None) -> str:
    output_path = _base_output('own_vs_competitors', output_path)
    df = _prepare_common(pd.DataFrame(list(records)), ['snapshot_date', 'competitor_name', 'entity_role', 'count', 'total_area', 'avg_price', 'total_price'])
    df = df.sort_values(['snapshot_date', 'entity_role', 'competitor_name'])

    raw = df[['snapshot_date', 'competitor_name', 'entity_role', 'count', 'total_area', 'avg_price', 'total_price']].rename(columns={'snapshot_date': 'Дата', 'competitor_name': 'Компания', 'entity_role': 'Роль', 'count': 'Помещений', 'total_area': 'Площадь, м²', 'avg_price': 'Средняя цена, ₽/м²', 'total_price': 'Суммарная стоимость, ₽'})
    area = _pivot(df, 'total_area', 'площадь, м²')
    count = _pivot(df, 'count', 'помещений')
    price = _pivot(df, 'avg_price', 'средняя цена, ₽/м²')
    total = _pivot(df, 'total_price', 'суммарная стоимость, ₽')
    latest = raw.loc[raw['Дата'] == str(df['snapshot_date'].max())].copy() if not df.empty else pd.DataFrame()

    try:
        from history_store import get_category_comparison_history
        category_records = get_category_comparison_history()
    except Exception:
        category_records = []

    commercial = _category_df(category_records, 'commercial')
    industrial = _category_df(category_records, 'industrial')
    commercial_area = _pivot(commercial, 'total_area', 'площадь, м²')
    commercial_count = _pivot(commercial, 'count', 'помещений')
    commercial_price = _pivot(commercial, 'avg_price', 'средняя цена, ₽/м²')
    industrial_area = _pivot(industrial, 'total_area', 'площадь, м²')
    industrial_count = _pivot(industrial, 'count', 'помещений')
    industrial_price = _pivot(industrial, 'avg_price', 'средняя цена, ₽/м²')

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        _write_sheet(writer, latest, 'Последний срез')
        _write_sheet(writer, raw, 'История по компаниям')
        _write_sheet(writer, area, 'Площадь')
        _write_sheet(writer, count, 'Помещения')
        _write_sheet(writer, price, 'Ставка')
        _write_sheet(writer, total, 'Стоимость')
        _write_sheet(writer, commercial_area, 'Офисы-торг площадь')
        _write_sheet(writer, commercial_count, 'Офисы-торг помещения')
        _write_sheet(writer, commercial_price, 'Офисы-торг ставка')
        _write_sheet(writer, industrial_area, 'Склады-пр-во площадь')
        _write_sheet(writer, industrial_count, 'Склады-пр-во помещения')
        _write_sheet(writer, industrial_price, 'Склады-пр-во ставка')
        _add_chart(writer, 'Площадь', 'Площадь по каждой компании', list(area.columns[1:]), 'Площадь, м²', 'A3')
        _add_chart(writer, 'Помещения', 'Количество помещений по каждой компании', list(count.columns[1:]), 'Количество', 'A27')
        _add_chart(writer, 'Ставка', 'Средняя ставка по каждой компании', list(price.columns[1:]), '₽/м²', 'A51')
        _add_chart(writer, 'Стоимость', 'Суммарная стоимость по каждой компании', list(total.columns[1:]), '₽', 'A75')
        _add_chart(writer, 'Офисы-торг площадь', 'Офисы / свободного назначения / торговые — площадь', list(commercial_area.columns[1:]), 'Площадь, м²', 'A3', 'Графики категорий')
        _add_chart(writer, 'Офисы-торг помещения', 'Офисы / свободного назначения / торговые — помещения', list(commercial_count.columns[1:]), 'Количество', 'A27', 'Графики категорий')
        _add_chart(writer, 'Офисы-торг ставка', 'Офисы / свободного назначения / торговые — ставка', list(commercial_price.columns[1:]), '₽/м²', 'A51', 'Графики категорий')
        _add_chart(writer, 'Склады-пр-во площадь', 'Производства / склады — площадь', list(industrial_area.columns[1:]), 'Площадь, м²', 'A75', 'Графики категорий')
        _add_chart(writer, 'Склады-пр-во помещения', 'Производства / склады — помещения', list(industrial_count.columns[1:]), 'Количество', 'A99', 'Графики категорий')
        _add_chart(writer, 'Склады-пр-во ставка', 'Производства / склады — ставка', list(industrial_price.columns[1:]), '₽/м²', 'A123', 'Графики категорий')
    return output_path
